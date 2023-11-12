#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@title: Bruker class
@description: methods and classes for the Bruker SolariX mass spectrometer 
@author: chrisarcadia 
@created: 2018/08/01
"""

import csv
import openpyxl
from lxml import etree
import sqlite3
import os
import numpy
import h5py
import copy
import time
import scipy.interpolate
from datetime import datetime

# function to write a picklist for the Solarix's MALDI modes

bruker_csv_fieldnames = ['Spot Number',
                      'Chip Number',
                      'Data Directory',
                      'Data File Name',
                      'Method Name',
                      'Sample Name',
                      'Comment',];
                         
def write_Bruker_xlsx_picklist(container,XLSXfilename,exclude_positions=[]):
        
    # create spreadsheet object
    now = datetime.now()
    workbook = openpyxl.Workbook();
    sheet = workbook.active
    sheet.title = "Target Automation Sample Table";
    
    # default field values
    chip = 0;
    directory = '';#'D:\\user_data\\Shared\\data\\';
    filename = '';#'Plate_X';
    method_name = '';#'D:\\Methods\\MS_Methods\\Install_180719\\MALDI.m';
    sample_name = ''; #container._description;
    sample_comment = ''; 
    
    # container info
    used_positions = container.list_filled_positions();             
    total_positions = container._rows*container._cols; 
    
    # exclude unwanted positions
    if exclude_positions != []:
        for pos in exclude_positions:
            if pos in used_positions:
                used_positions.remove(pos);
      
    # write headers
    fields = ['Spot Number',
              'Chip Number',
              'Data Directory',
              'Data File Name',
              'Method Name',
              'Sample Name',
              'Comment'];
    for n in range(0,len(fields)): 
        sheet.cell(row=1, column=n+1).value = fields[n];

    # write body
    counter = 1;
    for pos in used_positions: # Note: "Spot Number" format is X01Y01 for 1536 plate and A1 for 384 plate
        counter = counter + 1;
        if total_positions == 1536:
            spot_number = position_to_maldigrid(pos);
        elif total_positions == 384:
            spot_number = position_to_lettergrid(pos);
        else:
            raise ValueError('Unsupported format for plate type encountered when exporting a Bruker CSV picklist');        
        sheet.cell(row=counter, column=1).value = spot_number;
        sheet.cell(row=counter, column=2).value = chip;
        sheet.cell(row=counter, column=3).value = directory;
        sheet.cell(row=counter, column=4).value = filename;
        sheet.cell(row=counter, column=5).value = method_name;
        sheet.cell(row=counter, column=6).value = sample_name;
        sheet.cell(row=counter, column=7).value = sample_comment;
        
    # save spreadsheet
    workbook.save(filename = XLSXfilename);    
    print('Wrote Bruker XLSX picklist:',XLSXfilename)   
                         

def write_Bruker_csv_picklist(container,CSVfilename,label='',exclude_positions=[]):
    # could not get these CSV files to load in Solarix MALDI position loading window
    # *** use the XLSX method above   ***
    now = datetime.now()
    
    chip = 0;
    directory = 'D:\\user_data\\Shared\\data\\';
    filename = 'Plate_X';
    method_name = 'D:\Methods\MS_Methods\Install_180719\MALDI.m';
    sample_name = container._description + now.strftime("(%Y-%m-%d)");
    sample_comment = '';    
    if label == '':
        directory = directory + 'new';
    else:
        directory = directory + label;
        
    with open(CSVfilename, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=bruker_csv_fieldnames)    
        writer.writeheader()
        used_positions = container.list_filled_positions();       
        total_positions = container._rows*container._cols;  
        if exclude_positions != []:
            for pos in exclude_positions:
                used_positions.remove(pos);
        for pos in used_positions: # Note: "Spot Number" format is X01Y01 for 1536 plate and A1 for 384 plate
            if total_positions == 1536:
                spot_number = position_to_maldigrid(pos);
            elif total_positions == 384:
                spot_number = position_to_lettergrid(pos);
            else:
                raise ValueError('Unsupported format for plate type encountered when exporting a Bruker CSV picklist');
            writer.writerow({'Spot Number':spot_number,
                             'Chip Number':chip,
                             'Data Directory':directory,
                             'Data File Name':filename,
                             'Method Name':method_name,
                             'Sample Name':sample_name,
                             'Comment':sample_comment})

    print('Wrote Bruker CSV picklist:',CSVfilename)

def position_to_lettergrid(position):
    # Returns a string corresponding to the given (zero indexed) numerical position. Rows are letters and columns are numbers, such that (1,3) returns 'B4'.
    rowletters = [chr(x) for x in range(ord('A'),ord('Z')+1)] + ['A' + chr(x) for x in range(ord('A'),ord('Z')+1)];    
    if type(position)==tuple:
        return rowletters[position[0]]+str(position[1]+1)
    elif type(position)==list:
        return [position_to_lettergrid(x) for x in position]
    else:
        position = position; # not recognized. assume it does not need to be converted
    return position

def position_to_maldigrid(position):
    # Returns a string corresponding to the given (zero indexed) numerical position. Rows are indices of X and columns are indices of Y, such that (1,3) returns 'X01Y03'.    
    if type(position)==tuple:
        return 'X' + format(position[1]+1, '02d') + 'Y' + format(position[0]+1, '02d'); # X and Y are swapped in Bruker Solarix app 
    elif type(position)==list:
        return [position_to_maldigrid(x) for x in position]
    else:
        position = position; # not recognized. assume it does not need to be converted
    return position

# path navigation functions
    
def is_os_windows():
    windows = 'nt';
    return os.name == windows;

def ensure_os():
    if not is_os_windows():
        raise Exception('OS is not Windows')
    
def get_file_parts(filepath):
    # get the path, name, and extension of a file path
    path, name = os.path.split(filepath);
    name, extension = os.path.splitext(name);
    extension = extension[1:];
    return {'path':path,'name':name,'ext':extension};
    
def get_path_type(path):
    # returns path type 
    path_type = '';
    if os.path.splitext(path)[1]=='.d': # given path to a single data file
        path_type = 'data';
    elif os.path.isfile(path): # given path to an other sort of file
        path_type = 'file';
    elif os.path.isdir(path): # given path to a directory of multiple files
        path_type = 'dir';
    return path_type

def ensure_dir(path):
    os.path.exists(path)
    if not os.path.exists(path):
        os.makedirs(path);

def list_bruker_files(path):
    # returns full file names of all Bruker '.d' files in the specified directory
    file_list = [];
    path_type = get_path_type(path);
    if path_type=='data': 
        name = os.path.normpath(path);
        file_list.append(name);
    elif path_type=='dir': 
        for file in os.listdir(path):
            if get_path_type(file) == 'data':
                name = os.path.normpath(os.path.join(path,file));
                file_list.append(name);
    return file_list

def prep_output_dir(path_source,path_destin):
    # prepare the output directory
    path_source_type = get_path_type(path_source);
    if path_source_type == 'dir': # update destination with source folder name
        path_destin = path_destin + '\\'+ path_source.split('\\')[-1];
    ensure_dir(path_destin); # ensure destination path exists
    return path_destin;


# command line functions
    
def warn(message):
    print('Warning: ' + message)
    
# functions to read various database files
    
def import_XML(filename):
    # parse the XML file
    tree = etree.parse(filename);
    root = tree.getroot()
    return root;

def import_SQLite(filename):
    # initialize database tree
    tree = dict();

    # connect to database file
    conn = sqlite3.connect('file:' + filename +'?mode=ro', uri=True) # connect to database file in read only (ro) mode
    cursor = conn.cursor()
    
    # get table names
    cursor.execute("SELECT * FROM sqlite_master WHERE type='table';")
    result = cursor.fetchall()
    table_names = [el[2] for el in result];
    
    # get table data
    for name in table_names: 
        
            # get table headers
            cursor.execute("PRAGMA table_info('"+ name +"')");
            result = cursor.fetchall();
            column_names = [el[1] for el in result];
            body = dict();
            
            # get table body          
            for cname in column_names:
                cursor.execute('SELECT ' + cname + ' FROM ' + name + ' ORDER BY ' + column_names[0] + ' ASC');
                result = cursor.fetchall();
                data = [el[0] for el in result];
                body.update({cname:data});      
                
            # update database                   
            tree.update({name:body});
            
    # close connection to the database and return the imported data    
    conn.close() 
    return tree;

def import_NLDF(filename): # New Line Delimited File
    fid = open(filename);
    lines = [line.rstrip() for line in open(filename)];
    fid.close();
    return lines;

def import_NLDF_as_num(filename): # New Line Delimited File as Numerical List
    filedata = import_NLDF(filename);
    numerical_list = [];
    for line in filedata:
        try:
            numerical_list = numerical_list + [float(line)];
        except ValueError:
            numerical_list = []; # do nothing    return lines;
    return numerical_list;

# functions to help read measurement settings files directly (without CompassXtract)

def load_parameterChanges(filename):
    
    # Check if database exists
    DBfilename = os.path.normpath(filename + '\\' + 'parameterChanges.sqlite');
    if os.path.isfile(DBfilename):
    
        # load the SQlite file
        SQLite = import_SQLite(DBfilename);
        
        # get scan info
        scan_info = SQLite['ScanInfo'];
        scan_number = scan_info['ScanNumber'];
        scan_time = scan_info['TimeMinutes'];
        scan_TIC = scan_info['Tic'];
        scan_max_peak = scan_info['MaxPeak'];
        
        # get log of parameter changes 
        param_def = SQLite['ParameterDefinitions'];
        param_changes = SQLite['ParameterChanges']; 
        
        # get MALDI spot position (in order in which it they were measured)
        param_name = 'MALDI plate Spot number';
        if param_name in param_def['Description']:
            ind = param_def['Description'].index(param_name); # index of position parameter
            param_id = param_def['Id'][ind];
            log_scan_number = [];
            log_spot_number = [];
            for n in range(0,len(param_changes['ParameterId'])):
                if param_changes['ParameterId'][n] == param_id:
                    log_scan_number.append(param_changes['ScanNumber'][n]);
                    log_spot_number.append(param_changes['StringValue'][n].split(':')[0]);  
            scan_spot = []; 
            for n in range(0,len(scan_number)): # ensure spot number is in correct order
                number = scan_number[n];
                if number in log_scan_number:
                    ind = log_scan_number.index(number);
                    scan_spot.append(log_spot_number[ind]); 
        else:
            scan_spot = ['X01Y01']; # this is not a MALDI scan file so assume single position/spectra
            
        # format output
        result = {'maxPeak': scan_max_peak,
                  'number': scan_number,                  
                  'position': scan_spot,
                  'TIC': scan_TIC,
                  'elapsedMinutes': scan_time,};
    else:
        raise ValueError('File does not exist at: ' + DBfilename)
    return result

def load_method(filename):
    
    # Check if database exists
    methodfile = [path for path in os.listdir(filename) if path[-2:]=='.m'][0];
    DBfilename = os.path.normpath(filename + '\\' + methodfile + '\\' + 'apexAcquisition.method');
    if os.path.isfile(DBfilename):
    
        # load the XML file
        XML = import_XML(DBfilename);
        
        # get method name
        method_name = XML.find('methodmetadata').find('primarykey').find('methodname').text;
        method_date = XML.find('methodmetadata').find('primarykey').find('date').text.replace('_',' ');
        method_file_name = XML.find('methodmetadata').find('primarykey').find('methodfilepath').text;
        method_file_name_original = XML.find('methodmetadata').find('primarykey').find('methodloadfromfile').text;
        file_info = {'name': method_name, 
                     'date': method_date, 
                     'filename': method_file_name,
                     'filename_original': method_file_name_original};
                     
        # determine if file has MALDI info 
        el = XML.find('reportinfo');
        has_MALDI_info = el.find('.//section[@title="MALDI enabled"]') is not None;
        
        # helper function to retreive parameter values
        def retrieve(node,head,title,attribute):
            return parse_string_value(node.find('.//section[@title="'+ head +'"]//param[@title="' + title + '"]').attrib[attribute]);    

        # get ion info 
        el = XML.find('reportinfo');
        ion_info = {'polarity': retrieve(el,'Polarity','Polarity','value'),
                    'alternate_polarity': retrieve(el,'Polarity','Alternate Polarity','value'),};
        ion_info['ionization'] = 'MALDI';
        if not has_MALDI_info:
            ion_info['ionization'] = 'ESI';        

        # get acquisition info
        el = XML.find('reportinfo');
        mass_info = {'unit': retrieve(el,'Acquisition Mass Control','Broadband Low Mass','unit'),                                    
                     'low': retrieve(el,'Acquisition Mass Control','Broadband Low Mass','value'),                                    
                     'high': retrieve(el,'Acquisition Mass Control','Broadband High Mass','value'),                                                                      
                     'mode': retrieve(el,'Acquisition Mass Control','Detection Mode','value'),};
                     
        # get storage info                     
        storage_info =  {'perform_data_reduction': retrieve(el,'Data Storage','Perform Data Reduction','value'),  
                         'save_full_spectrum': retrieve(el,'Data Storage','Save Full Profile Spectrum','value'), 
                         'save_time_series': retrieve(el,'Data Storage','Save FID File','value'),
                         'acquisition_size': retrieve(el,'Acquisition Mass Control','Data Acquisition Size','value'), 
                         'processing_size': retrieve(el,'Acquisition Mass Control','Data Processing Size (SI)','value'),};

        # get time info
        el = XML.find('reportinfo'); 
        time_info = {'unit': retrieve(el,'Accumulation','Ion Accumulation Time','unit'),
                     'ion_accumulation': retrieve(el,'Accumulation','Ion Accumulation Time','value'),
                     'ion_cooling': retrieve(el,'Accumulation','Ion cooling time','value'),
                     'flight_to_detector': retrieve(el,'Accumulation','Time of Flight to Detector','value'),};

        # get chromatography info
        el = XML.find('reportinfo'); 
        chroma_info = {'LC_mode': retrieve(el,'Chromatography','LC Mode','value'),
                       'auto_MS_MS': retrieve(el,'Chromatography','Auto MS/MS','value'),
                       'ESI_high_voltage': retrieve(el,'Chromatography','ESI High Voltage','value'),
                       'source_quench': retrieve(el,'Chromatography','Source Quench','value'),};
                               
        # if available, get MALDI info
        if has_MALDI_info:
            # get laser info
            el = XML.find('reportinfo');
            laser_info = {'power_unit': retrieve(el,'MALDI','Laser Power','unit'),
                          'power': retrieve(el,'MALDI','Laser Power','value'), 
                          'shot_count': retrieve(el,'MALDI','Number of Laser Shots','value'),
                          'frequency': retrieve(el,'MALDI','Laser Frequency','value'),
                          'frequency_unit': retrieve(el,'MALDI','Laser Frequency','unit'),  
                          'focus': retrieve(el,'MALDI','LaserFocus','value'),};
                       
            # get smart walk info
            el = XML.find('reportinfo');
            walk_info = {'enabled': retrieve(el,'MALDI','MTP_SmartWalkEnable','value'),    
                         'pattern': retrieve(el,'MALDI','Smart walk pattern table','value'),        
                         'grid_width': retrieve(el,'MALDI','Smart walk grid width','value'),
                         'grid_width_unit': retrieve(el,'MALDI','Smart walk grid width','unit'),};  
                                  
            # get bias info
            el = XML.find('reportinfo');
            bias_info = {'unit': retrieve(el,'MALDI','Maldi Plate Offset','unit'),  
                         'MALDI_plate_offset': retrieve(el,'MALDI','Maldi Plate Offset','value'), 
                         'deflector_plate': retrieve(el,'MALDI','Deflector Plate','value'),};
        else:
            # get laser info
            laser_info = {};
                       
            # get smart walk info
            walk_info = {};  
                                  
            # get bias info
            bias_info = {};
                        

        # get (all) report info
        report_info = {};
        node = XML.find('reportinfo'); 
        for sec in node:
            sec_name = sec.attrib['title'];
            sec_dict = {};
            for subsec in sec:
                subsec_name = subsec.attrib['title'];
                subsec_dict = {};
                for param in subsec:
                    name = param.attrib['title'];
                    value = param.attrib['value'];
                    subsec_dict.update({parse_string_name(name): parse_string_value(value)})
                sec_dict.update({parse_string_name(subsec_name): subsec_dict})
            report_info.update({parse_string_name(sec_name): sec_dict})       

        # get parameter list info
        param_info = {};
        node = XML.find('paramlist'); 
        for el in node:
            value = parse_string_value(el.find('value').text);
            name = parse_string_name(el.attrib['name']);
            param_info.update({name:value})
            
        # get the sample period
        sample_rate = 2*param_info['SW_h'];
        storage_info.update({'sample_rate': sample_rate, 'sample_rate_unit': 'Hz'});
        
        # format output
        result = {'file': file_info,
                  'storage': storage_info,                  
                  'ion': ion_info,
                  'laser': laser_info,
                  'walk': walk_info,
                  'mass': mass_info,
                  'bias': bias_info,
                  'time': time_info,
                  'chroma': chroma_info,
                  'more': {'report': report_info,
                           'param': param_info,}
                  };
    else:
        raise ValueError('File does not exist at: ' + DBfilename)
    return result

def parse_string_value(string):
    value = string; # default value
    try: # converting to float
        if string is not None:
            value = float(string);
        else:
            value = '';
    except ValueError:
        value = string; # leave as string
    return value;

def parse_string_name(string):
    name = string; # default name
    name = name.replace(' ','_'); # replace spaces with underscore
    name = name[0:31]; # limit to 31 characters
    return name;

# functions to help read measurement data files (with CompassXtract)

def get_measurement_settings(filename):
    try: # grab measurement settings
        try: # grab scan settings
            scan_info = load_parameterChanges(filename);                
        except:
            error_source = 'scan';
            scan_info = {};
        try: # grab method settings
            method_info = load_method(filename);             
        except:
            error_source = 'method';
            method_info = {};
        settings = {'sample_rate':method_info['storage']['sample_rate'], # sample rate [Hz]
                    'acquisition_size': method_info['storage']['acquisition_size'], # samples per measurement
                    'positions': scan_info['position'], # measurements' positions on plate
                    'scan_info': scan_info,
                    'method_info': method_info,
                    };
        return settings;
    except:
        raise ValueError('Failed to load '+error_source+' settings from '+'"'+filename+'"');
    
def prep_conversion_file(filename):
    # prepare conversion output file
    if not os.path.isfile(filename):   
        path = get_file_parts(filename)['path'];
        if not os.path.isdir(path):
            os.makedirs(os.path.dirname(filename), exist_ok=True)        
        fileObject = h5py.File(filename, 'w');
        fileObject.close();           
        
def string_list_to_ASCII(strlist):
    return [el.encode("ascii", "ignore") for el in strlist]; # convert to ASCII for hdf5        
    
def convert_data(input_filename, output_filename, include_settings=True, include_spectra=True, include_raw=True, include_respectra=True, collect_sum=True, compression=None, copy_settings_to_mat=False, save_each_mz_vector=True, get_statistics=True):
    # main function for converting measurement data from '.d' file (in both the time and mass/charge domains)
    spectra_info = {};
    respectra_info = {};
    raw_info = {};
    # create new file
    if not os.path.isfile(output_filename):
        output_fileObject = h5py.File(output_filename, "w");
        name = get_file_parts(input_filename)['name'];
        currentdate = datetime.now();
        GMT_offset = time.strftime("+%H%M", time.gmtime(time.localtime().tm_gmtoff));
        date = currentdate.strftime("%a %b %m %H:%M:%S")+" GMT"+GMT_offset+" "+currentdate.strftime("%Y");        
        format_version = '1.0'; # converted file format version 
        format_name = 'aMALDIms'; # converted file format
        format_description = 'Arch MALDI Mass Spectrometry'; # converted file format description        
        format_documentation = 'https://github.com/jrosenstein/chemicalcpu/blob/master/chemcpupy/automation/Bruker/Setup.md';
        generator = 'Brown University'; # institution at which the data was generated
        header = {'filename': name,
                  'created': date,
                  'version': format_version,
                  'format': format_name,
                  'description': format_description,
                  'documentation': format_documentation,
                  'generator': generator,};
        output_fileObject.attrs.update(header)     
        hdf_options = output_fileObject.create_group("options");
        options = {'include_settings': include_settings,
                   'include_spectra': include_spectra,
                   'include_raw': include_raw,
                   'include_respectra': include_respectra,
                   'compression': str(compression),
                   'copy_settings_to_mat': copy_settings_to_mat,
                   'save_each_mz_vector' :save_each_mz_vector,
                   'collect_sum': collect_sum,                   
                   'get_statistics' : get_statistics};
        hdf_options.attrs.update(options);
        output_fileObject.close;
    else:
        raise ValueError('A file already exists at: ' + output_filename);
    # get measurement settings
    settings = get_measurement_settings(input_filename);
    # write settings to output file
    if include_settings:
        convert_settings(output_filename,settings,compression=compression);
    # write spectra to output file
    if include_spectra:
        spectra_info = convert_data_spectra(input_filename,output_filename,settings,compression=compression,save_each_mz_vector=save_each_mz_vector, get_statistics=get_statistics,collect_sum=False);
    # write resampled spectra to output file
    if include_respectra:
        respectra_info = resample_and_convert_data_spectra(input_filename,output_filename,settings,compression=compression,interpolation='cubic', get_statistics=get_statistics,collect_sum=collect_sum);
    # write raw to output file            
    if include_raw:
        raw_info = convert_data_raw(input_filename,output_filename,settings,compression=compression);  
    # put together settings
    info = {'header':header,
            'options':options,
            'settings':settings,
            'metadata':{'filename': output_filename,
                        'spectra': spectra_info, 
                        'respectra': respectra_info,
                        'raw': raw_info,}
            };
    # write  a '.mat' file with all found settings            
    if copy_settings_to_mat:
        export_settings_to_mat(output_filename,info);  
    return info;

def export_settings_to_mat(output_filename,all_settings):
        # save an exhaustive copy of file settings to a MAT (MATLAB file)
        import scipy.io
        fileparts = get_file_parts(output_filename);
        name = fileparts['name'];   
        path = fileparts['path'];
        description = 'This file contains all extracted parameters and settings but neither the raw or spectra data.';
        all_settings.update({'name': name,'description': description,})
        scipy.io.savemat(os.path.join(path,name+'.mat'), mdict=all_settings); # this will write the additional info to a '.mat' file

def convert_settings(output_filename, settings, compression=None):
    # method for converting data settings
    print('\t starting settings conversion')    
    output_fileObject = h5py.File(output_filename, "a");
    # check if fields already exists
    fields = list(output_fileObject.keys());
    # include scan info
    if ('scan' in fields):        
        warn('Failed to add scan log during settings conversion - scan field already present in file.')
    else:
        scan_info = copy.deepcopy(settings['scan_info']);
        hdf_scan = output_fileObject.create_group("scan");
        for name in scan_info.keys():  
            try:
                content = scan_info[name];
                content_has_strings = any(isinstance(el, str) for el in content);
                if content_has_strings:                    
                    content = string_list_to_ASCII(content); 
                hdf_scan.create_dataset(name, data=content,  compression=compression); 
            except:
                warn('Failed to include ' + name + ' from scan info.')
    # include method info
    if ('method' in fields):
        warn('Failed to add method info during settings conversion - method field already present in file.')
    else:        
        method_info = copy.deepcopy(settings['method_info']);
        hdf_method = output_fileObject.create_group("method");
        del method_info['more']; # remove more method details since hdf5 does not support nested dictionaries
        for name in method_info.keys(): 
            category = method_info[name];
            if isinstance(category,dict):
                hdf_category = hdf_method.create_group(name);
                hdf_category.attrs.update(category);
        output_fileObject.close();
    print('\t finished settings conversion')
    
def convert_data_raw(input_filename, output_filename, settings, label='raw', compression=None): 
    # method for converting raw data (time-domain) from '.d' file
    raw = {};
    bin_size = settings['acquisition_size'];
    bin_size = int(round(bin_size));    
    bin_bytes = bin_size * 4;  # int32 has 4 bytes    
    sample_rate = settings['sample_rate'];
    positions = settings['positions'];
    Npos = len(positions);
    Nsig = bin_size;
    # Check if raw data file exists (could be in one of two files depending on the number of measurements contained)
    filename1 = os.path.normpath(input_filename + '\\' + 'ser'); # a serial FID file with multiple raw measurements
    filename2 = os.path.normpath(input_filename + '\\' + 'fid'); # a FID file with a signle raw measurement
    found_file = False;
    if os.path.isfile(filename1):
        filename = filename1;
        found_file = True;
    elif os.path.isfile(filename2):
        filename = filename2;
        found_file = True;
    if found_file:
        print('\t starting raw conversion')
        # Open input and output files
        input_fileObject = open(filename, "r");
        prep_conversion_file(output_filename);
        output_fileObject = h5py.File(output_filename, "a");        
        # Check if fields already exists
        fields = list(output_fileObject.keys());
        if (label in fields):
            warn('Failed raw conversion - fields already present in file.')
        else:                    
            # Initialize the dataset
            hdf_raw = output_fileObject.create_group(label);
            print('\t \t initializing dataset')                
            hdf_signal = hdf_raw.create_dataset("signal", compression=compression, shape=(Npos,Nsig), dtype='int32');                                            
            # Loop the conversion             
            counter = 0;
            while True:    
                ser = numpy.fromfile(input_fileObject, dtype=numpy.int32, count=bin_size)
                if ser.size == 0:
                    break;  
                else:
                    counter = counter + 1;      
                    if counter<=Npos:
                        name = positions[counter-1];
                    else:
                        name = 'extra_'+str(counter-Npos);
                    hdf_signal[counter-1,0:Nsig] = ser;                    
                    print('\t \t raw #' + '{0:04d}'.format(counter) +  ' (' + name + ') converted')    
                    input_fileObject.seek(bin_bytes*counter, os.SEEK_SET)
                    output_fileObject.flush(); 
            if counter>0:
                time = numpy.divide(numpy.arange(0,bin_size,1),sample_rate);
                hdf_raw.create_dataset('position', data=string_list_to_ASCII(positions),  compression=compression);                 
                hdf_raw.create_dataset('time', data=time,  compression=compression); 
                info = {'count': counter,
                        'points': Nsig,
                        'duration': float(bin_size)/float(sample_rate),
                        'sample_rate':sample_rate,
                        'type':'int32',
                        'unit_duration':'s',
                        'unit_sample_rate':'Hz',
                       };
                hdf_raw.attrs.update(info)
                raw = info;
            else:
                warn('Failed raw conversion. File is empty.')   
            # Close input and output files               
            input_fileObject.close();
            output_fileObject.close();
        print('\t finished raw conversion')
    else:
        raise ValueError('File does not exist at: ' + input_filename)
    return raw;

def resample_and_convert_data_spectra(input_filename, output_filename, settings, label='respectra', compression=None, interpolation='cubic', get_statistics=True, collect_sum=True):
    # method for resampling and converting spectra data (mass/charge-domain) from '.d' file    
    respectra = convert_data_spectra(input_filename, output_filename, settings, label=label, compression=compression, save_each_mz_vector=False, syncronize_grid=True, syncronize_method=interpolation, get_statistics=get_statistics, collect_sum=collect_sum);
    return respectra;

def convert_data_spectra(input_filename, output_filename, settings, label='spectra', compression=None, save_each_mz_vector=True, syncronize_grid=False, syncronize_method='cubic', get_statistics=True, collect_sum=True):
    # method for converting spectra data (mass/charge-domain) from '.d' file
    # (spectrum type can be 'profile' for )
    spectra = {};                
    #positions = settings['positions'];
    #Npos = len(positions);
    filename = os.path.normpath(input_filename); 
    # Check if data file exists
    if os.path.isdir(filename): # use isdir since the '.d' file is a really a directory
        if is_os_windows():    
            print('\t starting spectra conversion')
            # Load Common Object Model (COM) library
            import comtypes.client  
            # Create CompassXtractMS object
            MSAnalysis = comtypes.client.CreateObject("EDAL.MSAnalysis"); # CompassXtractMS     
            MSAnalysis.Open(filename);
            name = MSAnalysis.AnalysisName;
            Nspec = len(MSAnalysis.MSSpectrumCollection); # number of spectra in file
            timestamp = MSAnalysis.AnalysisDateTimeIsoString; # time file was created [ISO 8601]
            param = get_MSAnalysis_parameters(MSAnalysis); # file parameters            
            # Check that there is data         
            if Nspec>0:
                # Find the index associated with spot number 
                spot_index = get_parameter(param,'Spot Number','index');                
                # Get available positions and summarize the file
                posDict = get_MSAnalysis_positions(MSAnalysis,spot_index);                
                Npos = len(posDict);
                positions = [];
                if not Nspec == Npos:
                    warn('Mismatch between number of positions (' + str(Npos) + ') and number of spectra (' + str(Nspec) + ') found in file.');        
                print('\t \t found ' +str(Nspec)+ ' unique spectra starting at ' + posDict[1]);
                # Open output files
                prep_conversion_file(output_filename);
                output_fileObject = h5py.File(output_filename, "a");
                # check if fields already exists
                fields = list(output_fileObject.keys());
                if (label in fields):
                    warn('Failed spectra conversion - fields already present in file.')
                else:
                    # Constants
                    polarities = [];
                    poles = {0:+1,1:-1,255:0}; # {0:'postive', 1:'negative', 255:'unknown'}  
                    calibrate = 1; # {0:'for raw data', 1: 'for recalibrated data'} # we should always use recalibrated                                            
                    # Get data size
                    spectrum = MSAnalysis.MSSpectrumCollection(1); # will get length of first spectrum
                    mzBase,intensity,points = spectrum.GetMassIntensityValues(calibrate);   
                    mzBase = numpy.array(mzBase, dtype = numpy.dtype(float));
                    Nsig = len(intensity);    
                    # Initialize the dataset
                    print('\t \t initializing dataset')                                    
                    hdf_spec = output_fileObject.create_group(label);                    
                    hdf_signal = hdf_spec.create_dataset("signal", compression=compression, shape=(Npos,Nsig), dtype='float');                    
                    # Loop the conversion
                    if save_each_mz_vector:
                        hdf_mz = hdf_spec.create_dataset("mass_to_charge", compression=compression, shape=(Npos,Nsig), dtype='f');                    
                    has_more_than_one_spectrum = Npos>1; 
                    if get_statistics:
                        stats = dict();
                        background = dict();
                        stats['max'] = numpy.zeros(Npos);
                        stats['min'] = numpy.zeros(Npos);
                        stats['mean'] = numpy.zeros(Npos);
                        stats['stdev'] = numpy.zeros(Npos); 
                        stats['median'] = numpy.zeros(Npos);
                        stats['quartile1'] = numpy.zeros(Npos);
                        stats['quartile3'] = numpy.zeros(Npos);  
                        stats['sum'] = numpy.zeros(Npos);
                        stats['area'] = numpy.zeros(Npos);                        
                        background['noiseGauss3Sigma'] = numpy.zeros(Npos); 
                        background['noiseGauss6Sigma'] = numpy.zeros(Npos); 
                        background['noiseTukeyInner'] = numpy.zeros(Npos); 
                        background['noiseTukeyOuter'] = numpy.zeros(Npos); 
                        background['offsetGauss3Sigma'] = numpy.zeros(Npos); 
                        background['offsetGauss6Sigma'] = numpy.zeros(Npos); 
                        background['offsetTukeyInner'] = numpy.zeros(Npos); 
                        background['offsetTukeyOuter'] = numpy.zeros(Npos);                         
                        background['cutoffGauss3Sigma'] = numpy.zeros(Npos); 
                        background['cutoffGauss6Sigma'] = numpy.zeros(Npos); 
                        background['cutoffTukeyInner'] = numpy.zeros(Npos); 
                        background['cutoffTukeyOuter'] = numpy.zeros(Npos);   
                        background['peakcountGauss3Sigma'] = numpy.zeros(Npos); 
                        background['peakcountGauss6Sigma'] = numpy.zeros(Npos); 
                        background['peakcountTukeyInner'] = numpy.zeros(Npos); 
                        background['peakcountTukeyOuter'] = numpy.zeros(Npos); 
                        get_offset = lambda intensity, threshold : numpy.mean(intensity[intensity<threshold]); # get the mean of the background signal (that which is below the threshold)                        
                        get_noise = lambda intensity, threshold : numpy.std(intensity[intensity<threshold]); # get the standard deviation of the background signal (that which is below the threshold)
                        get_peakcount = lambda intensity, threshold : int(numpy.floor(numpy.sum(numpy.abs(numpy.diff(intensity<threshold)))/2)); # get the number of peaks above threshold
                        peaklevel = 9; # currently using 3 times (2 times was not quick high enough to avoid noise peaks for large mz) the 3-sigma empirical rule # number of noise levels (background standard deviations) a peak must be greater than        
                    if collect_sum and syncronize_grid and has_more_than_one_spectrum:
                        intensity_resampled_sum = numpy.zeros((1,Nsig));
                    for n in range(Npos):
                        spectrum = MSAnalysis.MSSpectrumCollection(n + 1);
                        mz,intensity,points = spectrum.GetMassIntensityValues(calibrate); 
                        intensity = numpy.array(intensity, dtype = numpy.dtype(float));
                        mz = numpy.array(mz, dtype = numpy.dtype(float));
                        if not len(intensity)==Nsig:
                            warn('Spectrum length not equal to expected length ('+str(Nsig)+').');
                        if syncronize_grid and has_more_than_one_spectrum:
                            intensity_resampled = scipy.interpolate.griddata(mz, intensity, mzBase, method=syncronize_method); # available interpolation methods: {‘linear’, ‘nearest’, ‘cubic’}
                            hdf_signal[n,0:Nsig] = intensity_resampled;
                            if collect_sum:
                                intensity_resampled_sum = intensity_resampled_sum + intensity_resampled;                            
                        else:
                            hdf_signal[n,0:Nsig] = intensity;
                        if save_each_mz_vector:
                            hdf_mz[n,0:Nsig] =  mz;                             
                        name = posDict[n + 1];                        
                        positions.append(name);   
                        # collect spectrum polarity
                        polarity = spectrum.Polarity;
                        if polarity in poles:
                            polarity = poles[polarity]
                        polarities.append(polarity);
                        output_fileObject.flush(); 
                        # collect intensity statistics
                        if get_statistics:
                            stats['max'][n] = numpy.max(intensity);
                            stats['min'][n] = numpy.min(intensity);
                            stats['mean'][n] = numpy.mean(intensity);
                            stats['stdev'][n] = numpy.std(intensity);
                            stats['median'][n] = numpy.median(intensity); # 50th percentile                           
                            stats['quartile1'][n] = numpy.percentile(intensity,25); # 25th percentile
                            stats['quartile3'][n] = numpy.percentile(intensity,75); # 75th percentile
                            stats['sum'][n] = numpy.sum(intensity);   
                            stats['area'][n] = numpy.trapz(intensity,x=mz); # area under the curve                              
                            Gauss3Sigma = stats['mean'][n] + 3*stats['stdev'][n]; # Gaussian distribution empirical rule (for below +3 standard deviations)                   
                            Gauss6Sigma = stats['mean'][n] + 6*stats['stdev'][n]; # Gaussian distribution empirical rule (for below +6 standard deviations)                                                                           
                            TukeyInner = stats['quartile3'][n] + 1.5*(stats['quartile3'][n]-stats['quartile1'][n]); # Tukey inside upper fence (below +1.5 hinge steps)     
                            TukeyOuter = stats['quartile3'][n] + 3*(stats['quartile3'][n]-stats['quartile1'][n]); # Tukey outside upper fence (below +3 hinge steps)                                 
                            background['noiseGauss3Sigma'][n] = get_noise(intensity,Gauss3Sigma); # background noise level below Gaussian empirical rule 
                            background['noiseGauss6Sigma'][n] = get_noise(intensity,Gauss6Sigma); # background noise level below twice the Gaussian empirical rule   
                            background['noiseTukeyInner'][n] = get_noise(intensity,TukeyInner); # background noise level below upper bound of Tukey inside fence                                              
                            background['noiseTukeyOuter'][n] = get_noise(intensity,TukeyOuter); # background noise level below upper bound of Tukey outside fence                                        
                            background['offsetGauss3Sigma'][n] = get_offset(intensity,Gauss3Sigma); # background offset level below Gaussian empirical rule 
                            background['offsetGauss6Sigma'][n] = get_offset(intensity,Gauss6Sigma); # background offset level below twice the Gaussian empirical rule   
                            background['offsetTukeyInner'][n] = get_offset(intensity,TukeyInner); # background offset level below upper bound of Tukey inside fence                                              
                            background['offsetTukeyOuter'][n] = get_offset(intensity,TukeyOuter); # background offset level below upper bound of Tukey outside fence                                        
                            background['cutoffGauss3Sigma'][n] = Gauss3Sigma; 
                            background['cutoffGauss6Sigma'][n] = Gauss6Sigma; 
                            background['cutoffTukeyInner'][n] = TukeyInner; 
                            background['cutoffTukeyOuter'][n] = TukeyOuter;   
                            background['peakcountGauss3Sigma'][n] = get_peakcount(intensity,background['offsetGauss3Sigma'][n]+peaklevel*background['noiseGauss3Sigma'][n]); 
                            background['peakcountGauss6Sigma'][n] = get_peakcount(intensity,background['offsetGauss6Sigma'][n]+peaklevel*background['noiseGauss6Sigma'][n]); 
                            background['peakcountTukeyInner'][n] = get_peakcount(intensity,background['offsetTukeyInner'][n]+peaklevel*background['noiseTukeyInner'][n]); 
                            background['peakcountTukeyOuter'][n] = get_peakcount(intensity,background['offsetTukeyOuter'][n]+peaklevel*background['noiseTukeyOuter'][n]);                
# replaced with above code (2018/12/21)                                                                                                                                                                          
#                            background['peakcountGauss3Sigma'][n] = get_peakcount(intensity,peaklevel*Gauss3Sigma); 
#                            background['peakcountGauss6Sigma'][n] = get_peakcount(intensity,peaklevel*Gauss6Sigma); 
#                            background['peakcountTukeyInner'][n] = get_peakcount(intensity,peaklevel*TukeyInner); 
#                            background['peakcountTukeyOuter'][n] = get_peakcount(intensity,peaklevel*TukeyOuter);                                                                                                            
                        print('\t \t spectra #' + '{0:04d}'.format(n+1) +  ' (' + name + ') converted')    
                    if len(mz)>0:     
                        hdf_spec.create_dataset('position', data=string_list_to_ASCII(positions),  compression=compression);                         
                        if not save_each_mz_vector:
                            hdf_spec.create_dataset('mass_to_charge', data=mzBase,  compression=compression); # currently using knowledge that all mz series are the same for all spectra in a single file      
                        all_same_polarity = all(pol == polarity for pol in polarities);
                        if not all_same_polarity:
                            hdf_spec.create_dataset('polarity', data=polarities, compression=compression);
                            polarity = 0; 
                        hdf_param = hdf_spec.create_group("parameters");        
                        hdf_param.attrs.update(format_MSAnalysis_parameters_for_HDF5(param));
                        if get_statistics:
                            hdf_stat = hdf_spec.create_group("statistics");  
                            for key in stats.keys():
                                hdf_stat.create_dataset(key, data=stats[key],  compression=compression);
                            hdf_background = hdf_spec.create_group("background");  
                            for key in background.keys():
                                hdf_background.create_dataset(key, data=background[key],  compression=compression);
                            hdf_background.attrs.update({'peak_level':peaklevel});                                                         
                        info = {'count': Npos,
                                'points': points,
                                'created': timestamp,
                                'ion_polarity': polarity,
                               };
                        if syncronize_grid:
                            info.update({'interpolant': syncronize_method});
                            if collect_sum and has_more_than_one_spectrum:
                                hdf_spec.create_dataset('signal_sum', data=intensity_resampled_sum, compression=compression, dtype='float');                                
                        hdf_spec.attrs.update(info)
                        spectra = info;
                        spectra.update({'parameters':format_MSAnalysis_parameters_for_MATLAB(param)});
                    else:
                        warn('Failed spectra conversion. Spectrum was empty.')         
                # Close output files               
                output_fileObject.close();    
            else:
                warn('Failed spectra conversion. File is empty.')      
            print('\t finished spectra conversion')
        else:        
            warn('Failed spectra conversion. Requires Windows OS')            
    else:
        raise ValueError('File does not exist at: ' + filename)
    return spectra;

def get_parameter(parameters,parameter,field):
    if parameter in parameters:
        return parameters[parameter][field];
    else:
        raise ValueError('Missing "'+parameter+'" parameter in serial file.');   

    
def get_MSAnalysis_parameters(MSAnalysis,spectrum_index=1):
    # get spectrum parameters
    spectrum = MSAnalysis.MSSpectrumCollection(spectrum_index); # grab first spectrum by default
    parameters = dict();
    n = 0;
    while True: # loop until we reach an index out of bounds error
        try: 
            n = n+1
            param = spectrum.MSSpectrumParameterCollection(n);
            name = param.ParameterName;
            value = param.ParameterValue;
            #unit = p.ParameterUnit if p.ParameterUnit is not None else "";
            parameters.update({name:{'value':value,'index':n}});
        except: 
            break
    return parameters;
    # note: for easy reading, we can show the keys in order with: sorted(list(param))
    
def format_MSAnalysis_parameters_for_HDF5(parameters):
    parametersHDF5 = dict();
    for name in parameters:
        entry = parameters[name];
        value = entry['value'];
        parametersHDF5.update({name:value});
    return parametersHDF5
    
def format_MSAnalysis_parameters_for_MATLAB(parameters):
    parametersMATLAB = dict();
    for name in parameters:
        entry = parameters[name];
        entry['name'] = name;        
        key = name.replace("[Instrument Param]","").replace(" ","_").replace("-","_");
        key = "".join([i for i in key if i.isalpha() or i=="_"])
        key = key[0:31]; # limit the length of the key
        parametersMATLAB.update({key:entry});
    return parametersMATLAB
                
def get_MSAnalysis_positions(MSAnalysis,spot_index):
    # get spot position/location of all spectra
    positions = {};
    positions_unique = {};           
    name = MSAnalysis.AnalysisName;
    i = 1;
    while True:
        try:
            spectrum = MSAnalysis.MSSpectrumCollection(i);
            param = spectrum.MSSpectrumParameterCollection(spot_index);
            stringID = param.ParameterValue.partition(':')[0];
            positions[i] = stringID; # generate a dictionary of well IDs                                 
            i = i + 1
        except:        
            # Check for double recordings of the same well in the serial
            for key,value in positions.items():
                if value not in positions_unique.values():
                    positions_unique[key] = value
            Ndup = len(positions) - len(positions_unique);
            if Ndup>=1:
                warn(str(name) + ' contains ' +str(Ndup)+ ' duplicate spectra: ')         
                print(str(set(positions.items()).symmetric_difference(set(positions_unique.items()))))
            return positions # return all positions (including the duplicates)
            break
